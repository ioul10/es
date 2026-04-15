"""
core/pdf_to_excel.py
Extraction PDF fiscal → Excel structuré.

Principe : on laisse pdfplumber extraire les tableaux tels quels (version de base),
puis on remet en forme chaque feuille avec un cadre propre et les bonnes colonnes.

Deux formats détectés automatiquement :
  - AMMC (5 pages) : Actif, Passif, CPC répartis sur pages 2-5
  - DGI  (7 pages) : Actif (p2-3), Passif (p4), CPC (p5-7)
"""

import re
from collections import defaultdict
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logger import get_logger

logger = get_logger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
C_DARK   = "1F3864"
C_MED    = "2E75B6"
C_LIGHT  = "D6E4F0"
C_WHITE  = "FFFFFF"
C_GRAY   = "F5F7FA"
C_BORDER = "B8CCE4"
NUM_FMT  = '#,##0.00;[Red]-#,##0.00;"-"'

# ── Styles ────────────────────────────────────────────────────────────────────

def _b():
    s = Side(style='thin', color=C_BORDER)
    return Border(top=s, bottom=s, left=s, right=s)

def _cell(ws, r, c, v=None, bg=C_WHITE, fg="222222", bold=False,
          align="left", num_fmt=None, sz=9, wrap=True, indent=0):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", size=sz, bold=bold, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap, indent=indent)
    cell.border    = _b()
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def _row_style(label: str):
    """Retourne (bg, fg, bold) selon le label."""
    n = str(label).strip().upper()
    if re.match(r'^TOTAL', n):
        return C_DARK, C_WHITE, True
    if re.match(r'^R[EÉ]SULTAT', n):
        return "2E4057", C_WHITE, True
    # Ligne section : tout en majuscules > 4 caractères alphanumériques
    alph = re.sub(r'[^A-Z0-9]', '', n)
    if len(alph) > 4 and alph == re.sub(r'[^A-Z0-9]', '', str(label).upper()):
        return C_LIGHT, C_DARK, True
    return C_WHITE, "333333", False

# ── Utilitaires ───────────────────────────────────────────────────────────────

def _is_rotated(v) -> bool:
    """Lettres rotatives A\nC\nT\nI\nF."""
    if not v: return False
    parts = [p.strip() for p in str(v).split('\n') if p.strip()]
    return len(parts) >= 3 and all(len(p) <= 2 and p.replace('.','').isalpha() for p in parts)

def _parse(s) -> float | None:
    """Parse nombre FR : '1 234 567,89' ou float déjà."""
    if s is None: return None
    if isinstance(s, (int, float)): return float(s)
    s = str(s).strip().replace('\xa0','').replace(' ','')
    if not s or s in ['-','—','/']: return None
    neg = s.startswith('-')
    s = s.lstrip('-')
    m = re.match(r'^(\d{1,3}(?:\.\d{3})*),(\d{2})$', s)
    if m: s = m.group(1).replace('.','') + '.' + m.group(2)
    elif re.match(r'^\d+,\d{2}$', s): s = s.replace(',','.')
    elif re.match(r'^\d+(\.\d+)?$', s): pass
    else: return None
    try: return -float(s) if neg else float(s)
    except: return None

def _clean(v) -> str:
    """Nettoyage cellule : retirer lettres rotatives, uniformiser espaces."""
    if v is None: return ''
    s = str(v).strip()
    if _is_rotated(s): return ''
    return re.sub(r'\s+', ' ', re.sub(r'\n+', ' ', s)).strip()

def _is_skip_row(row: list) -> bool:
    """Lignes à ignorer : indices 0 1 2 3, ou headers PDF, ou lignes vides."""
    vals = [str(c).strip() for c in row if c is not None]
    if not vals: return True
    # Indices numériques croissants 0 1 2 3 ...
    try:
        nums = [int(v) for v in vals]
        if nums == list(range(len(nums))): return True
    except (ValueError, TypeError): pass
    return False

# ── Extraction infos générales ─────────────────────────────────────────────────

def _extract_info(pdf) -> dict:
    info = {}
    for i in range(min(2, len(pdf.pages))):
        text = pdf.pages[i].extract_text() or ''
        for key, pat in [
            ('raison_sociale',      r'[Rr]aison\s+[Ss]ociale\s*:?\s*([A-Z][^\n]{3,60})'),
            ('identifiant_fiscal',  r'[Ii]dentifiant\s+[Ff]iscal\s*:?\s*(\d+)'),
            ('taxe_professionnelle',r'[Tt]axe\s+[Pp]rof\w*\.?\s*:?\s*([\d\s]+)'),
            ('adresse',             r'[Aa]dresse\s*:?\s*([^\n]{5,60})'),
        ]:
            if key not in info:
                m = re.search(pat, text, re.IGNORECASE)
                if m: info[key] = m.group(1).strip()

        if 'exercice' not in info:
            for pat in [
                r'p[eé]riode\s+du\s+(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})',
                r'[Ee]xercice\s+du\s+(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})',
                r'(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})',
            ]:
                m = re.search(pat, text)
                if m:
                    info['exercice']       = f"Du {m.group(1)} au {m.group(2)}"
                    info['exercice_fin']   = m.group(2)
                    break

    for k in ('raison_sociale','identifiant_fiscal','taxe_professionnelle',
              'adresse','exercice','exercice_fin'):
        info.setdefault(k, '')
    return info


# ── Détection cellules fusionnées + fallback X/Y ──────────────────────────────

def _has_fused(table) -> bool:
    """Tableau avec plusieurs valeurs dans une cellule (\n)."""
    count = 0
    for row in (table or []):
        for cell in (row or []):
            if cell and '\n' in str(cell):
                parts = [x for x in str(cell).split('\n') if _parse(x) is not None]
                if len(parts) > 1:
                    count += 1
    return count >= 3


def _xy_rows(page) -> list:
    """
    Extraction X/Y pour pages avec cellules fusionnées (SGTM, etc.).
    Retourne une liste de pseudo-lignes [section, label, v1, v2, v3, v4].
    """
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    if not words: return []

    num_ws = [w for w in words
              if re.match(r'^-?\d+$', w['text'].replace(',','').replace('.',''))
              and len(w['text'].replace(',','').replace('.','')) >= 1
              and w['x0'] > 100]
    if not num_ws: return []
    thresh = min(w['x0'] for w in num_ws) - 5

    lines = defaultdict(list)
    for w in words:
        lines[round(w['top'] / 3) * 3].append(w)

    result = []
    for y in sorted(lines):
        row = sorted(lines[y], key=lambda w: w['x0'])
        lw = [w for w in row if w['x0'] < thresh]
        nw = [w for w in row if w['x0'] >= thresh
              and re.match(r'^-?\d+$', w['text'].replace(',','').replace('.',''))]

        # Label
        filt = [w for w in lw
                if not (len(w['text']) <= 1
                        and re.match(r'^[A-Z.]$', w['text'])
                        and w['x0'] < 50)]
        label = ''
        if filt:
            label = filt[0]['text']
            for i in range(1, len(filt)):
                gap = filt[i]['x0'] - filt[i-1]['x1']
                label += filt[i]['text'] if gap <= 1 else ' ' + filt[i]['text']
            label = re.sub(r' +', ' ', label).strip()

        # Valeurs
        vals = []
        if nw:
            grp = [nw[0]]
            for w in nw[1:]:
                if w['x0'] - grp[-1]['x1'] < 18:
                    grp.append(w)
                else:
                    v = _parse(''.join(x['text'] for x in grp))
                    if v is not None: vals.append(v)
                    grp = [w]
            v = _parse(''.join(x['text'] for x in grp))
            if v is not None: vals.append(v)

        if label and vals:
            result.append(['', label] + [str(v) for v in vals])
        elif label and not vals:
            # Label sans valeurs : garder pour rattachement éventuel
            result.append(['', label])
        elif not label and vals and result:
            # Valeurs orphelines → rattacher au dernier label sans valeurs
            last = result[-1]
            if len(last) == 2:   # label seul
                result[-1] = last + [str(v) for v in vals]
            # sinon ignorer (valeurs appartiennent à une ligne déjà complète)

    # Filtrer les lignes sans valeurs
    return [r for r in result if len(r) > 2]

# ── Bloc titre commun ──────────────────────────────────────────────────────────

def _write_title_block(ws, title: str, info: dict, n_cols: int) -> int:
    """Écrit les 4 premières lignes (titre + infos). Retourne la prochaine ligne."""
    raison   = info.get('raison_sociale','')
    if_num   = info.get('identifiant_fiscal','')
    exercice = info.get('exercice','')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    _cell(ws, 1, 1, title, bg=C_DARK, fg=C_WHITE, bold=True, align='center', sz=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1,n_cols-1))
    _cell(ws, 2, 1, f"Raison sociale : {raison}", bg=C_LIGHT, fg=C_DARK,
          bold=True, sz=9, indent=1)
    _cell(ws, 2, n_cols, f"IF : {if_num}", bg=C_LIGHT, fg=C_DARK, align='right', sz=9)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    _cell(ws, 3, 1, f"Exercice : {exercice}", bg=C_GRAY, fg="555555", sz=9, indent=1)

    ws.row_dimensions[4].height = 4  # séparateur
    return 5  # prochaine ligne disponible

# ── Feuille Identification ─────────────────────────────────────────────────────

def _write_ident(wb, info: dict):
    ws = wb.create_sheet("1 - Identification")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 52

    ws.merge_cells('A1:B1')
    _cell(ws, 1, 1, "PIÈCES ANNEXES À LA DÉCLARATION FISCALE",
          bg=C_DARK, fg=C_WHITE, bold=True, align='center', sz=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:B2')
    _cell(ws, 2, 1, "IMPÔTS SUR LES SOCIÉTÉS — Modèle Comptable Normal (loi 9-88)",
          bg=C_MED, fg=C_WHITE, align='center', sz=10)
    ws.row_dimensions[2].height = 18

    fields = [
        ("Raison sociale",        info.get('raison_sociale','—')),
        ("Identifiant fiscal",    info.get('identifiant_fiscal','—')),
        ("Taxe professionnelle",  info.get('taxe_professionnelle','—')),
        ("Adresse",               info.get('adresse','—')),
        ("Date de bilan",         info.get('exercice_fin','—')),
    ]
    for i, (lbl, val) in enumerate(fields, 4):
        ws.row_dimensions[i].height = 18
        _cell(ws, i, 1, lbl, bg=C_LIGHT, fg=C_DARK, bold=True, sz=9, indent=1)
        _cell(ws, i, 2, val, bg=C_WHITE,  fg="222222", sz=9, indent=1)

# ── Extraction + écriture Bilan Actif ─────────────────────────────────────────
#
# PDF AMMC : Col1=lettre_rot, Col2=Label, Col3=Brut, Col4=vide, Col5=Amort, Col6=NetN, Col7=NetN1
# PDF DGI  : Col1=section,   Col2=Label, Col3=Brut, Col4=Amort, Col5=NetN, Col6=NetN1
#
# Excel cible : ColA=Label | ColB=Brut | ColC=Amort | ColD=Net(N) | ColE=Net(N-1)

def _detect_val_cols(table) -> list:
    """Détecte les colonnes numériques (1-based) par fréquence d'apparition."""
    hits = {}
    for row in (table or [])[2:]:
        if not row: continue
        for ci, c in enumerate(row):
            if not c: continue
            s = str(c).strip().replace('\xa0','').replace(' ','')
            if re.match(r'^-?\d{1,3}(\.\d{3})*,\d{2}$', s) or re.match(r'^-?\d+,\d{2}$', s):
                hits[ci+1] = hits.get(ci+1, 0) + 1
    return sorted(k for k, v in hits.items() if v >= 2)


def _write_actif(wb, info: dict, tables: list, n_pdf_cols: int):
    """
    Construit la feuille Bilan Actif.
    Détecte dynamiquement les colonnes numériques de chaque tableau.
    """
    ws = wb.create_sheet("2 - Bilan Actif")
    ws.sheet_view.showGridLines = False

    n_cols = 5  # Label + 4 valeurs
    r = _write_title_block(ws, "BILAN ACTIF", info, n_cols)

    col_defs = [
        ("DÉSIGNATION",         50),
        ("BRUT",                18),
        ("AMORT. & PROV.",      18),
        ("NET — EXERCICE N",    18),
        ("NET — EXERCICE N-1",  18),
    ]
    for ci, (h, w) in enumerate(col_defs, 1):
        _cell(ws, r, ci, h, bg=C_MED, fg=C_WHITE, bold=True, align='center', sz=9, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 28
    ws.freeze_panes = f'A{r+1}'
    r += 1

    written = 0
    seen = set()

    for table in tables:
        # Détecter si c'est un pseudo-tableau X/Y (col0='', col1=label, col2+=float)
        is_xy = (table and len(table) > 0 and table[0] and
                 str(table[0][0]).strip() == '' and len(table[0]) >= 3 and
                 _parse(table[0][2]) is not None)

        if is_xy:
            # X/Y : col0='', col1=label, col2=v1, col3=v2, col4=v3, col5=v4
            for row in table:
                if not row or len(row) < 2: continue
                label = _clean(row[1] if len(row) > 1 else None)
                if not label or len(label) < 2: continue
                if _parse(label) is not None: continue
                key = re.sub(r'\W', '', label.lower())
                if key in seen: continue
                seen.add(key)
                # Valeurs à partir de col2
                vals_xy = [_parse(row[i]) for i in range(2, min(6, len(row)))]
                n = len(vals_xy)
                if n >= 4:
                    brut, amort, netN, netN1 = vals_xy[0], vals_xy[1], vals_xy[2], vals_xy[3]
                elif n == 3:
                    brut, amort, netN, netN1 = vals_xy[0], None, vals_xy[1], vals_xy[2]
                elif n == 2:
                    brut, amort, netN, netN1 = vals_xy[0], None, None, vals_xy[1]
                elif n == 1:
                    brut, amort, netN, netN1 = vals_xy[0], None, None, None
                else:
                    brut = amort = netN = netN1 = None
                bg, fg, bold = _row_style(label)
                ws.row_dimensions[r].height = 15 if bg == C_WHITE else 17
                indent = 1 if bg == C_WHITE else 0
                _cell(ws, r, 1, label, bg=bg, fg=fg, bold=bold, align='left', sz=9, indent=indent)
                for ci, v in enumerate([brut, amort, netN, netN1], 2):
                    _cell(ws, r, ci, 0 if v is None else v,
                          bg=bg, fg=fg, bold=bold, align='right', sz=9, num_fmt=NUM_FMT)
                r += 1
                written += 1
        else:
            # Tableau standard : détecter colonnes numériques
            val_cols = _detect_val_cols(table)

            for row in table:
                if not row or _is_skip_row(row): continue
                if len(row) < 2: continue

                label = _clean(row[1] if len(row) > 1 else None)
                if not label or len(label) < 2: continue
                if _parse(label) is not None: continue

                key = re.sub(r'\W', '', label.lower())
                if key in seen: continue
                seen.add(key)

                def gv(col):
                    if col and len(row) >= col: return _parse(row[col-1])
                    return None

                n = len(val_cols)
                if n >= 4:
                    brut, amort, netN, netN1 = gv(val_cols[0]), gv(val_cols[1]), gv(val_cols[2]), gv(val_cols[3])
                elif n == 3:
                    brut, amort, netN, netN1 = gv(val_cols[0]), None, gv(val_cols[1]), gv(val_cols[2])
                elif n == 2:
                    brut, amort, netN, netN1 = gv(val_cols[0]), None, None, gv(val_cols[1])
                elif n == 1:
                    brut, amort, netN, netN1 = gv(val_cols[0]), None, None, None
                else:
                    brut = amort = netN = netN1 = None

                bg, fg, bold = _row_style(label)
                ws.row_dimensions[r].height = 15 if bg == C_WHITE else 17
                indent = 1 if bg == C_WHITE else 0

                _cell(ws, r, 1, label, bg=bg, fg=fg, bold=bold, align='left', sz=9, indent=indent)
                for ci, v in enumerate([brut, amort, netN, netN1], 2):
                    _cell(ws, r, ci, 0 if v is None else v,
                          bg=bg, fg=fg, bold=bold, align='right', sz=9, num_fmt=NUM_FMT)
                r += 1
                written += 1

    return written

# ── Extraction + écriture Bilan Passif ────────────────────────────────────────
#
# PDF AMMC : Col1=lettre_rot, Col2=Label, Col3=vide, Col4=ExN, Col5=ExN1
# PDF DGI  : Col1=section,   Col2=Label, Col3=ExN,  Col4=ExN1
#
# Excel cible : ColA=Label | ColB=Exercice N | ColC=Exercice N-1

def _write_passif(wb, info: dict, tables: list, n_pdf_cols: int):
    ws = wb.create_sheet("3 - Bilan Passif")
    ws.sheet_view.showGridLines = False

    n_cols = 3
    r = _write_title_block(ws, "BILAN PASSIF", info, n_cols)

    col_defs = [
        ("DÉSIGNATION",     54),
        ("EXERCICE N",      20),
        ("EXERCICE N-1",    20),
    ]
    for ci, (h, w) in enumerate(col_defs, 1):
        _cell(ws, r, ci, h, bg=C_MED, fg=C_WHITE, bold=True, align='center', sz=9, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 28
    ws.freeze_panes = f'A{r+1}'
    r += 1

    written = 0
    seen = set()

    for table in tables:
        is_xy = (table and len(table) > 0 and table[0] and
                 str(table[0][0]).strip() == '' and len(table[0]) >= 3 and
                 _parse(table[0][2]) is not None)

        for row in table:
            if not row or _is_skip_row(row): continue
            if len(row) < 2: continue

            label = _clean(row[1] if len(row) > 1 else None)
            if not label or len(label) < 2: continue
            if _parse(label) is not None: continue

            key = re.sub(r'\W', '', label.lower())
            if key in seen: continue
            seen.add(key)

            if is_xy:
                vals_xy = [_parse(row[i]) for i in range(2, min(4, len(row)))]
                exN  = vals_xy[0] if len(vals_xy) > 0 else None
                exN1 = vals_xy[1] if len(vals_xy) > 1 else None
            else:
                val_cols = _detect_val_cols(table)
                def get_val(col):
                    if col and len(row) >= col: return _parse(row[col-1])
                    return None
                n = len(val_cols)
                exN  = get_val(val_cols[0]) if n >= 1 else None
                exN1 = get_val(val_cols[1]) if n >= 2 else None

            bg, fg, bold = _row_style(label)
            ws.row_dimensions[r].height = 15 if bg == C_WHITE else 17
            indent = 1 if bg == C_WHITE else 0

            _cell(ws, r, 1, label, bg=bg, fg=fg, bold=bold, align='left', sz=9, indent=indent)
            for ci, v in enumerate([exN, exN1], 2):
                _cell(ws, r, ci, 0 if v is None else v,
                      bg=bg, fg=fg, bold=bold, align='right', sz=9, num_fmt=NUM_FMT)
            r += 1
            written += 1

    # Note légale
    r2 = r + 1
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=3)
    c = ws.cell(r2, 1)
    c.value = "(1) Capital personnel débiteur.  (2) Bénéficiaire (+) / Déficitaire (−)."
    c.font  = Font(name="Arial", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return written

# ── Extraction + écriture CPC ─────────────────────────────────────────────────
#
# PDF AMMC : Col1=lettre_rot, Col2=Numéro, Col3=Label, Col4=PropresN,
#            Col5=ExercPrec, Col6=TotalN, Col7=TotalN1
# PDF DGI  : Col1=section, Col2=Label, Col3=PropresN, Col4=ExercPrec,
#            Col5=TotalN, Col6=TotalN1  (sur 3 pages fusionnées)
#
# Excel cible : ColA=Num | ColB=Label | ColC=PropresN | ColD=ExercPréc | ColE=TotalN | ColF=TotalN1

ROMAIN_RE = re.compile(
    r'^(I{1,3}|IV|V?I{0,3}|IX|XI{0,3}|XIV|XV|XVI)\.?\s*$', re.I)

def _write_cpc(wb, info: dict, tables: list, n_pdf_cols: int):
    ws = wb.create_sheet("4 - CPC")
    ws.sheet_view.showGridLines = False

    n_cols = 6
    r = _write_title_block(ws, "COMPTE DE PRODUITS ET CHARGES (Hors Taxes)", info, n_cols)

    col_defs = [
        ("N°",                     5),
        ("DÉSIGNATION",            42),
        ("PROPRES À\nL'EXERCICE",  18),
        ("EXERCICES\nPRÉCÉDENTS",  18),
        ("TOTAUX\nEXERCICE N",     18),
        ("TOTAUX\nEXERCICE N-1",   18),
    ]
    for ci, (h, w) in enumerate(col_defs, 1):
        _cell(ws, r, ci, h, bg=C_MED, fg=C_WHITE, bold=True, align='center', sz=9, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 30
    ws.freeze_panes = f'A{r+1}'
    r += 1

    written = 0
    seen = set()

    for table in tables:
        is_xy = (table and len(table) > 0 and table[0] and
                 str(table[0][0]).strip() == '' and len(table[0]) >= 3 and
                 _parse(table[0][2]) is not None)
        val_cols = [] if is_xy else _detect_val_cols(table)

        for row in table:
            if not row or _is_skip_row(row): continue
            if len(row) < 2: continue

            num_raw = _clean(row[0] if row else None)
            num = num_raw if (num_raw and ROMAIN_RE.match(num_raw)) else ''

            label = _clean(row[1] if len(row) > 1 else None)
            if not label or len(label) < 2: continue
            if _parse(label) is not None: continue

            key = re.sub(r'\W', '', label.lower())
            if key in seen: continue
            seen.add(key)

            if is_xy:
                v_list = [_parse(row[i]) for i in range(2, min(6, len(row)))]
                propre = v_list[0] if len(v_list) > 0 else None
                prec   = v_list[1] if len(v_list) > 1 else None
                totN   = v_list[2] if len(v_list) > 2 else None
                totN1  = v_list[3] if len(v_list) > 3 else None
            else:
                def get_val(col):
                    if col and len(row) >= col: return _parse(row[col-1])
                    return None
                n = len(val_cols)
                propre = get_val(val_cols[0]) if n >= 1 else None
                prec   = get_val(val_cols[1]) if n >= 2 else None
                totN   = get_val(val_cols[2]) if n >= 3 else None
                totN1  = get_val(val_cols[3]) if n >= 4 else None

            bg, fg, bold = _row_style(label)
            ws.row_dimensions[r].height = 15 if bg == C_WHITE else 17
            indent = 1 if bg == C_WHITE else 0

            _cell(ws, r, 1, num or None, bg=bg,
                  fg=fg if bg != C_WHITE else C_MED,
                  bold=True, align='center', sz=8)
            _cell(ws, r, 2, label, bg=bg, fg=fg, bold=bold,
                  align='left', sz=9, indent=indent)
            for ci, v in enumerate([propre, prec, totN, totN1], 3):
                _cell(ws, r, ci, 0 if v is None else v,
                      bg=bg, fg=fg, bold=bold, align='right', sz=9, num_fmt=NUM_FMT)
            r += 1
            written += 1

    # Notes
    r2 = r + 1
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=6)
    c = ws.cell(r2, 1)
    c.value = ("(1) Stock final − Stock initial.  "
               "(2) Achats revendus = Achats − Variation de stock.")
    c.font  = Font(name="Arial", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return written

# ── Point d'entrée ────────────────────────────────────────────────────────────

def convert(pdf_path: str, output_path: str) -> dict:
    """
    Convertit un PDF fiscal → Excel structuré.
    Détecte automatiquement AMMC (5 pages) ou DGI (7 pages).
    """
    pdf     = pdfplumber.open(pdf_path)
    n_pages = len(pdf.pages)
    info    = _extract_info(pdf)

    is_dgi = (n_pages == 7)
    fmt    = "DGI" if is_dgi else "AMMC"
    logger.info(f"Format : {fmt} ({n_pages} pages)")

    # Mapping pages (index 0-based)
    if is_dgi:
        actif_pages  = [1, 2]    # pages 2-3
        passif_pages = [3]        # page 4
        cpc_pages    = [4, 5, 6]  # pages 5-7
    else:
        actif_pages  = [1]        # page 2
        passif_pages = [2]        # page 3
        cpc_pages    = [3, 4]     # pages 4-5

    # Extraire les tableaux bruts de chaque section
    def get_tables(indices):
        tables = []
        for idx in indices:
            if idx >= len(pdf.pages): continue
            page = pdf.pages[idx]
            raw = page.extract_tables()
            good = [t for t in raw if t and len(t) >= 2 and t[0] and len(t[0]) >= 2]
            if not good:
                continue
            if any(_has_fused(t) for t in good):
                logger.info(f"  page {idx+1} → X/Y (cellules fusionnées)")
                xy = _xy_rows(page)
                if xy:
                    tables.append(xy)
            else:
                tables.extend(good)
        return tables

    actif_tables  = get_tables(actif_pages)
    passif_tables = get_tables(passif_pages)
    cpc_tables    = get_tables(cpc_pages)
    pdf.close()

    # Déterminer le nb de colonnes source
    def n_cols(tables):
        for t in tables:
            for row in t:
                if row:
                    return len(row)
        return 7

    nc_actif  = n_cols(actif_tables)
    nc_passif = n_cols(passif_tables)
    nc_cpc    = n_cols(cpc_tables)
    logger.info(f"Colonnes PDF → actif:{nc_actif} passif:{nc_passif} cpc:{nc_cpc}")

    # Construire l'Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_ident(wb, info)
    n_a = _write_actif(wb,  info, actif_tables,  nc_actif)
    n_p = _write_passif(wb, info, passif_tables, nc_passif)
    n_c = _write_cpc(wb,    info, cpc_tables,    nc_cpc)

    wb.save(output_path)
    total = n_a + n_p + n_c
    logger.info(f"Sauvegardé : {total} lignes ({n_a}a / {n_p}p / {n_c}c)")

    return {
        'tables':      3,
        'rows':        total,
        'pages':       n_pages,
        'format':      fmt,
        'info':        info,
        'exercice_fin': info.get('exercice_fin', ''),
    }
