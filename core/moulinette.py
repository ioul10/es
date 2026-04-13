"""
core/moulinette.py
Remplit la moulinette standard à partir de 1 à 3 fichiers Excel FiscalXL.
D = N-2 (plus ancien), E = N-1, F = N (plus récent)
"""
import openpyxl, re, unicodedata, shutil
from pathlib import Path

# ── Template moulinette (chemin fixe dans le projet) ─────────────────────────
MOL_TEMPLATE = Path(__file__).parent.parent / "mol_v0.xlsx"

# ══════════════════════════════════════════════════════════════════
# MAPPING : (row_moulinette, keyword_fiscalxl, index_dans_vals)
# actif  vals = [Brut, Amort, NetN, NetN1]   → NetN  = index 2
# passif vals = [ExN, ExN1]                  → ExN   = index 0
# cpc    vals = [PropN, ExPrec, TotN, TotN1] → TotN  = index 2
# ══════════════════════════════════════════════════════════════════

ACTIF_MAP = [
    (9,  'frais preliminaires',                  2),
    (10, 'charges repartir',                     2),
    (11, 'primes remboursement',                 2),
    (13, 'immobilisations recherche',            2),
    (14, 'brevets marques droits',               2),
    (15, 'fonds commercial',                     2),
    (16, 'autres immobilisations incorporelles', 2),
    (18, 'terrains',                             2),
    (19, 'constructions',                        2),
    (20, 'installations techniques',             2),
    (21, 'materiel transport',                   2),
    (22, 'mobilier materiel bureau',             2),
    (23, 'autres immobilisations corporelles',   2),
    (24, 'immobilisations corporelles cours',    2),
    (26, 'prets immobilises',                    2),
    (27, 'autres creances financieres',          2),
    (28, 'titres participation',                 2),
    (29, 'autres titres immobilises',            2),
    (31, 'diminution creances immobilisees',     2),
    (32, 'augmentations dettes financement',     2),
    (35, 'marchandises',                         2),
    (36, 'matieres fournitures consommables',    2),
    (37, 'produits cours',                       2),
    (38, 'produits intermediaires residuels',    2),
    (39, 'produits finis',                       2),
    (41, 'fournisseurs debiteurs avances',       2),
    (42, 'clients comptes rattaches',            2),
    (43, 'personnel actif circulant',            2),
    (44, 'etat actif',                           2),
    (45, 'comptes associes actif',               2),
    (46, 'autres debiteurs',                     2),
    (47, 'comptes regularisation actif',         2),
    (48, 'titres valeurs placement',             2),
    (49, 'ecarts conversion actif circulant',    2),
    (52, 'cheques valeurs encaisser',            2),
    (53, 'banques tg ccp',                       2),
    (54, 'caisse regie avances',                 2),
]

PASSIF_MAP = [
    (10, 'capital social personnel',          0),
    (11, 'prime emission fusion',             0),
    (12, 'ecarts reevaluation',               0),
    (13, 'reserve legale',                    0),
    (14, 'autres reserves',                   0),
    (15, 'report nouveau',                    0),
    (16, 'resultat instance affectation',     0),
    (17, 'resultat net exercice',             0),
    (19, 'subvention investissement',         0),
    (20, 'provisions reglementees',           0),
    (22, 'emprunts obligataires',             0),
    (23, 'autres dettes financement',         0),
    (26, 'provisions risques',                0),
    (27, 'provisions charges',                0),
    (30, 'fournisseurs comptes rattaches',    0),
    (31, 'clients crediteurs avances',        0),
    (32, 'personnel passif',                  0),
    (33, 'organismes sociaux',                0),
    (34, 'etat passif',                       0),
    (35, 'comptes associes passif',           0),
    (36, 'autres creanciers',                 0),
    (37, 'comptes regularisation passif',     0),
    (38, 'autres provisions risques charges', 0),
    (39, 'ecarts conversion passif',          0),
    (42, 'credits escompte',                  0),
    (43, 'credits tresorerie',                0),
    (44, 'banques soldes crediteurs',         0),
]

CPC_MAP = [
    (10, 'ventes marchandises',                2),
    (11, 'ventes biens services',              2),
    (13, 'variation stocks produits',          2),
    (14, 'immobilisations produites',          2),
    (15, 'subventions exploitation',           2),
    (16, 'autres produits exploitation',       2),
    (17, 'reprises exploitation',              2),
    (20, 'achats revendus marchandises',       2),
    (21, 'achats consommes matieres',          2),
    (22, 'autres charges externes',            2),
    (23, 'impots taxes',                       2),
    (24, 'charges personnel',                  2),
    (25, 'autres charges exploitation',        2),
    (26, 'dotations exploitation',             2),
    (30, 'produits titres participation',      2),
    (31, 'gains change',                       2),
    (32, 'interets autres produits financiers',2),
    (33, 'reprises financieres',               2),
    (36, 'charges interets',                   2),
    (37, 'pertes change',                      2),
    (38, 'autres charges financieres',         2),
    (39, 'dotations financieres',              2),
    (44, 'produits cessions immobilisations',  2),
    (45, 'subventions equilibre',              2),
    (46, 'reprises subventions investissement',2),
    (47, 'autres produits non courants',       2),
    (48, 'reprises non courantes',             2),
    (51, 'valeurs nettes amortissements',      2),
    (53, 'autres charges non courantes',       2),
    (54, 'dotations non courantes',            2),
    (58, 'impots resultats',                   2),
]


# ══════════════════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════════════════

def _norm(s: str) -> str:
    s = unicodedata.normalize('NFD', str(s or ''))
    s = s.encode('ascii', 'ignore').decode().lower()
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def _get_year(exercice_str: str) -> int:
    m = re.findall(r'\d{4}', str(exercice_str or ''))
    return int(m[-1]) if m else 0


def _find_value(data: dict, keyword: str, col_idx: int):
    """Recherche fuzzy du mot-clé dans data, retourne la valeur à col_idx."""
    kwords = set(w for w in _norm(keyword).split() if len(w) > 2)
    best_score, best_val = 0, None
    for label_n, vals in data.items():
        lwords = set(w for w in label_n.split() if len(w) > 2)
        if not lwords:
            continue
        common = len(kwords & lwords)
        union  = len(kwords | lwords)
        score  = common / union if union > 0 else 0
        if score > best_score and score >= 0.35:
            best_score = score
            best_val   = vals[col_idx] if col_idx < len(vals) else None
    return best_val


# ══════════════════════════════════════════════════════════════════
# LECTURE FISCALXL
# ══════════════════════════════════════════════════════════════════

def read_fiscalxl(path: str) -> dict:
    """
    Lit un fichier Excel FiscalXL et retourne ses données structurées.
    Supporte les feuilles nommées '1 - Identification', '2 - Bilan Actif', etc.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    exercice = ''
    societe  = ''

    # Identification
    if '1 - Identification' in wb.sheetnames:
        wi = wb['1 - Identification']
        for r in range(1, wi.max_row + 1):
            lbl = str(wi.cell(r, 1).value or '').lower()
            val = str(wi.cell(r, 2).value or '')
            if 'exercice' in lbl and not exercice:
                exercice = val
            if ('soci' in lbl or 'raison' in lbl) and not societe:
                societe = val

    def _read_sheet(sheet_name, val_cols):
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        data = {}
        for r in range(6, ws.max_row + 1):
            lbl = ws.cell(r, 1).value
            if not lbl:
                continue
            n = _norm(lbl)
            if len(n) < 3:
                continue
            data[n] = [ws.cell(r, c).value for c in val_cols]
        return data

    return {
        'exercice': exercice,
        'societe':  societe,
        'annee':    _get_year(exercice),
        'actif':    _read_sheet('2 - Bilan Actif',  [2, 3, 4, 5]),
        'passif':   _read_sheet('3 - Bilan Passif', [2, 3]),
        'cpc':      _read_sheet('4 - CPC',          [2, 3, 4, 5]),
    }


# ══════════════════════════════════════════════════════════════════
# REMPLISSAGE MOULINETTE
# ══════════════════════════════════════════════════════════════════

def fill_moulinette(fiscalxl_paths: list, output_path: str) -> dict:
    """
    Remplit la moulinette standard avec les données des fichiers FiscalXL.
    Trie automatiquement par année : D=N-2, E=N-1, F=N.

    Args:
        fiscalxl_paths : liste de 1 à 3 chemins vers des Excel FiscalXL
        output_path    : chemin du fichier moulinette remplie à générer

    Returns:
        dict avec 'filled' (nb cellules), 'datasets' (info par exercice)
    """
    if not fiscalxl_paths:
        raise ValueError("Au moins 1 fichier FiscalXL requis.")
    if len(fiscalxl_paths) > 3:
        raise ValueError("Maximum 3 fichiers FiscalXL (3 exercices).")

    # Lire et trier par année (du plus ancien au plus récent)
    datasets = [read_fiscalxl(p) for p in fiscalxl_paths]
    datasets.sort(key=lambda d: d['annee'])

    # Copier le template moulinette
    shutil.copy(str(MOL_TEMPLATE), output_path)
    wb = openpyxl.load_workbook(output_path)

    # D=col4 (N-2), E=col5 (N-1), F=col6 (N)
    COLS = [4, 5, 6]
    filled = 0

    def _inject(ws, row_map, section_key):
        nonlocal filled
        for ds_idx, dataset in enumerate(datasets):
            col_excel = COLS[ds_idx]
            for row_mol, keyword, val_idx in row_map:
                v = _find_value(dataset[section_key], keyword, val_idx)
                if v is not None:
                    ws.cell(row_mol, col_excel).value = v
                    filled += 1

    _inject(wb['Feuil1'], ACTIF_MAP,  'actif')
    _inject(wb['Feuil2'], PASSIF_MAP, 'passif')
    _inject(wb['Feuil3'], CPC_MAP,    'cpc')

    # En-têtes dates dans les colonnes D/E/F
    for ds_idx, d in enumerate(datasets):
        col   = COLS[ds_idx]
        label = f"31/12/{d['annee']}\n(MAD)"
        wb['Feuil1'].cell(7, col).value = label
        wb['Feuil2'].cell(8, col).value = label

    wb.save(output_path)

    return {
        'filled':   filled,
        'datasets': [
            {
                'societe':  d['societe'],
                'exercice': d['exercice'],
                'annee':    d['annee'],
                'col':      ['D (N-2)', 'E (N-1)', 'F (N)'][i],
            }
            for i, d in enumerate(datasets)
        ],
    }
