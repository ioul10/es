"""
sgtm_excel_writer.py — Génération Excel pour la liasse fiscale SGTM
Structure Excel identique aux autres formats FiscalXL :
  Feuille 1 : Identification
  Feuille 2 : Bilan Actif     (49 lignes fixes)
  Feuille 3 : Bilan Passif    (44 lignes fixes)
  Feuille 4 : CPC             (54 lignes fixes)

Compatible avec la fonction write() attendue par app.py :
  stats = write(parsed, output_path)
  stats → {format, rows, actif, passif, cpc}
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from utils.logger import get_logger

logger = get_logger(__name__)

# ── Palette couleurs FiscalXL ──────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # Bleu foncé header
C_HEADER_FG   = "FFFFFF"   # Blanc
C_TOTAL_BG    = "BDD7EE"   # Bleu clair totaux
C_SUBTOTAL_BG = "DEEAF1"   # Bleu très clair sous-totaux
C_SECTION_BG  = "2E75B6"   # Bleu section
C_SECTION_FG  = "FFFFFF"
C_ALT_BG      = "F8FBFE"   # Alternance légère
C_BORDER      = "9DC3E6"

NUM_FMT = '#,##0.00'  # Format nombre marocain adapté

# ── Helpers style ──────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=9):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin"):
    s = Side(style=style, color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _header_row(ws, row, values, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, height=18):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(bg)
        cell.font = _font(bold=bold, color=fg, size=9)
        cell.alignment = _align("center", "center", wrap=True)
        cell.border = _border()


def _data_row(ws, row, libelle, val_n, val_n1,
              is_total=False, is_section=False, is_subtotal=False,
              indent=0, extra_cols=None):
    """
    Écrit une ligne de données.
    extra_cols : liste de valeurs supplémentaires (ex: Brut, Amort pour l'actif)
    """
    ws.row_dimensions[row].height = 14

    if is_section:
        bg, fg, bold = C_SECTION_BG, C_SECTION_FG, True
    elif is_total:
        bg, fg, bold = C_TOTAL_BG, "1F3864", True
    elif is_subtotal:
        bg, fg, bold = C_SUBTOTAL_BG, "1F3864", True
    else:
        bg, fg, bold = (C_ALT_BG if row % 2 == 0 else "FFFFFF"), "000000", False

    # Col 1 : libellé
    lbl_cell = ws.cell(row=row, column=1, value=(" " * indent) + libelle)
    lbl_cell.fill  = _fill(bg)
    lbl_cell.font  = _font(bold=bold, color=fg)
    lbl_cell.alignment = _align("left", "center", wrap=True)
    lbl_cell.border = _border()

    # Colonnes numériques
    col_idx = 2
    all_vals = (extra_cols or []) + [val_n, val_n1]
    for v in all_vals:
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.fill   = _fill(bg)
        cell.font   = _font(bold=bold, color=fg)
        cell.alignment = _align("right", "center")
        cell.border  = _border()
        if isinstance(v, (int, float)):
            cell.number_format = NUM_FMT
        col_idx += 1


# ── Feuille 1 : Identification ─────────────────────────────────────────────────

def _write_identification(wb, info: dict):
    ws = wb.create_sheet("1 - Identification")
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, 'A', 35)
    _set_col_width(ws, 'B', 55)

    # Titre
    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value = "LIASSE FISCALE — IDENTIFICATION"
    c.fill  = _fill(C_HEADER_BG)
    c.font  = _font(bold=True, color=C_HEADER_FG, size=12)
    c.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 25

    fields = [
        ("Raison Sociale",       info.get("raison_sociale", "")),
        ("Identifiant Fiscal",   info.get("identifiant_fiscal", "")),
        ("Taxe Professionnelle", info.get("taxe_pro", "")),
        ("Adresse",              info.get("adresse", "")),
        ("Date de Déclaration",  info.get("date_declaration", "")),
        ("Fin Exercice",         info.get("exercice_fin", "")),
        ("Format Liasse",        "SGTM — Modèle Normal IS (5 pages)"),
    ]

    for i, (label, value) in enumerate(fields, 2):
        ws.row_dimensions[i].height = 18
        cl = ws.cell(row=i, column=1, value=label)
        cl.fill = _fill(C_SUBTOTAL_BG)
        cl.font = _font(bold=True, color="1F3864")
        cl.alignment = _align("left", "center")
        cl.border = _border()

        cv = ws.cell(row=i, column=2, value=value)
        cv.fill = _fill("FFFFFF")
        cv.font = _font()
        cv.alignment = _align("left", "center")
        cv.border = _border()


# ── Feuille 2 : Bilan Actif ────────────────────────────────────────────────────

def _write_actif(wb, actif: dict, exercice_fin: str) -> int:
    ws = wb.create_sheet("2 - Bilan Actif")
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, 'A', 48)
    _set_col_width(ws, 'B', 18)
    _set_col_width(ws, 'C', 18)
    _set_col_width(ws, 'D', 18)
    _set_col_width(ws, 'E', 18)

    a = actif
    row = 1

    # En-tête
    _header_row(ws, row,
        ["BILAN ACTIF — MCN loi 9-88",
         f"Brut {exercice_fin}", "Amortissements & Provisions",
         f"Net {exercice_fin}", "Net Exercice Précédent"],
        height=22)
    row += 1

    def R(libelle, brut, net_n, net_n1,
          amort=None, is_total=False, is_section=False,
          is_subtotal=False, indent=0):
        nonlocal row
        _data_row(ws, row, libelle, net_n, net_n1,
                  is_total=is_total, is_section=is_section,
                  is_subtotal=is_subtotal, indent=indent,
                  extra_cols=[brut, amort if amort is not None else (brut - net_n)])
        row += 1

    def amort_calc(brut, net):
        return max(0.0, brut - net)

    # ── ACTIF IMMOBILISÉ ────────────────────────────────────────────────────
    R("A — ACTIF IMMOBILISÉ", 0, 0, 0, is_section=True)

    R("Immobilisations en non-valeurs",
      a.get('nv_total_brut',0), a.get('nv_total_net',0), a.get('nv_total_net1',0),
      is_subtotal=True)
    R("  Frais préliminaires",
      a.get('nv_frais_prelim_brut',0), a.get('nv_frais_prelim_net',0),
      a.get('nv_frais_prelim_net1',0), indent=2)
    R("  Charges à répartir sur plusieurs exercices",
      a.get('nv_charges_repartir_brut',0), a.get('nv_charges_repartir_net',0),
      a.get('nv_charges_repartir_net1',0), indent=2)
    R("  Prime de remboursement des obligations",
      a.get('nv_prime_remboursement_brut',0), a.get('nv_prime_remboursement_net',0),
      a.get('nv_prime_remboursement_net1',0), indent=2)

    R("Immobilisations incorporelles",
      a.get('inc_total_brut',0), a.get('inc_total_net',0), a.get('inc_total_net1',0),
      is_subtotal=True)
    R("  Immobilisations en R&D",
      a.get('inc_rd_brut',0), a.get('inc_rd_net',0), a.get('inc_rd_net1',0), indent=2)
    R("  Brevets, marques, droits et valeurs similaires",
      a.get('inc_brevets_brut',0), a.get('inc_brevets_net',0), a.get('inc_brevets_net1',0), indent=2)
    R("  Fonds commercial",
      a.get('inc_fonds_commercial_brut',0), a.get('inc_fonds_commercial_net',0),
      a.get('inc_fonds_commercial_net1',0), indent=2)
    R("  Autres immobilisations incorporelles",
      0, 0, 0, indent=2)

    R("Immobilisations corporelles",
      a.get('corp_total_brut',0), a.get('corp_total_net',0), a.get('corp_total_net1',0),
      is_subtotal=True)
    R("  Terrains",
      a.get('corp_terrains_brut',0), a.get('corp_terrains_net',0), a.get('corp_terrains_net1',0), indent=2)
    R("  Constructions",
      a.get('corp_constructions_brut',0), a.get('corp_constructions_net',0),
      a.get('corp_constructions_net1',0), indent=2)
    R("  Installations techniques, matériel et outillage",
      a.get('corp_instal_brut',0), a.get('corp_instal_net',0), a.get('corp_instal_net1',0), indent=2)
    R("  Matériel de transport",
      a.get('corp_mater_transport_brut',0), a.get('corp_mater_transport_net',0),
      a.get('corp_mater_transport_net1',0), indent=2)
    R("  Mobilier, matériel de bureau et aménagements divers",
      a.get('corp_mobilier_brut',0), a.get('corp_mobilier_net',0), a.get('corp_mobilier_net1',0), indent=2)
    R("  Autres immobilisations corporelles",
      a.get('corp_autres_brut',0), a.get('corp_autres_net',0), a.get('corp_autres_net1',0), indent=2)
    R("  Immobilisations corporelles en cours",
      a.get('corp_en_cours_brut',0), a.get('corp_en_cours_net',0), a.get('corp_en_cours_net1',0), indent=2)

    R("Immobilisations financières",
      a.get('fin_total_brut',0), a.get('fin_total_net',0), a.get('fin_total_net1',0),
      is_subtotal=True)
    R("  Prêts immobilisés",
      a.get('fin_prets_brut',0), a.get('fin_prets_net',0), a.get('fin_prets_net1',0), indent=2)
    R("  Autres créances financières", 0, 0, 0, indent=2)
    R("  Titres de participation",
      a.get('fin_titres_participation_brut',0), a.get('fin_titres_participation_net',0),
      a.get('fin_titres_participation_net1',0), indent=2)
    R("  Autres titres immobilisés",
      a.get('fin_autres_titres_brut',0), a.get('fin_autres_titres_net',0),
      a.get('fin_autres_titres_net1',0), indent=2)

    R("Écarts de conversion — Actif (éléments durables)",
      a.get('eca_immo_total_brut',0), a.get('eca_immo_total_net',0),
      a.get('eca_immo_total_net1',0), is_subtotal=True)

    R("TOTAL I — ACTIF IMMOBILISÉ",
      a.get('total_i_brut',0), a.get('total_i_net',0), a.get('total_i_net1',0),
      is_total=True)

    # ── ACTIF CIRCULANT ─────────────────────────────────────────────────────
    R("B — ACTIF CIRCULANT", 0, 0, 0, is_section=True)

    R("Stocks",
      a.get('stocks_total_brut',0), a.get('stocks_total_net',0), a.get('stocks_total_net1',0),
      is_subtotal=True)
    R("  Marchandises",
      a.get('stocks_marchandises_brut',0), a.get('stocks_marchandises_net',0),
      a.get('stocks_marchandises_net1',0), indent=2)
    R("  Matières et fournitures consommables",
      a.get('stocks_matieres_brut',0), a.get('stocks_matieres_net',0),
      a.get('stocks_matieres_net1',0), indent=2)
    R("  Produits en cours", 0, 0, 0, indent=2)
    R("  Produits intermédiaires et résiduels", 0, 0, 0, indent=2)
    R("  Produits finis",
      a.get('stocks_produits_finis_brut',0), a.get('stocks_produits_finis_net',0),
      a.get('stocks_produits_finis_net1',0), indent=2)

    R("Créances de l'actif circulant",
      a.get('cre_ac_total_brut',0), a.get('cre_ac_total_net',0), a.get('cre_ac_total_net1',0),
      is_subtotal=True)
    R("  Fournisseurs débiteurs, avances et acomptes",
      a.get('cre_ac_fournisseurs_brut',0), a.get('cre_ac_fournisseurs_net',0),
      a.get('cre_ac_fournisseurs_net1',0), indent=2)
    R("  Clients et comptes rattachés",
      a.get('cre_ac_clients_brut',0), a.get('cre_ac_clients_net',0),
      a.get('cre_ac_clients_net1',0), indent=2)
    R("  Personnel", 0, 0, 0, indent=2)
    R("  État",
      a.get('cre_ac_etat_brut',0), a.get('cre_ac_etat_net',0), a.get('cre_ac_etat_net1',0), indent=2)
    R("  Comptes d'associés", 0, 0, 0, indent=2)
    R("  Autres débiteurs",
      a.get('cre_ac_autres_brut',0), a.get('cre_ac_autres_net',0), a.get('cre_ac_autres_net1',0), indent=2)
    R("  Comptes de régularisation — Actif", 0, 0, 0, indent=2)

    R("Titres et valeurs de placement",
      a.get('tvp_total_brut',0), a.get('tvp_total_net',0), a.get('tvp_total_net1',0),
      is_subtotal=True)

    R("Écarts de conversion — Actif circulant",
      a.get('eca_ac_total_brut',0), a.get('eca_ac_total_net',0), a.get('eca_ac_total_net1',0),
      is_subtotal=True)

    R("TOTAL II — ACTIF CIRCULANT",
      a.get('total_ii_brut',0), a.get('total_ii_net',0), a.get('total_ii_net1',0),
      is_total=True)

    # ── TRÉSORERIE ACTIF ────────────────────────────────────────────────────
    R("C — TRÉSORERIE ACTIF", 0, 0, 0, is_section=True)

    R("Chèques et valeurs à encaisser",
      a.get('treso_cheques_brut',0), a.get('treso_cheques_net',0), a.get('treso_cheques_net1',0), indent=2)
    R("Banques, T.G. et C.C.P.",
      a.get('treso_banque_brut',0), a.get('treso_banque_net',0), a.get('treso_banque_net1',0), indent=2)
    R("Caisses, régies d'avances et accréditifs",
      a.get('treso_caisses_brut',0), a.get('treso_caisses_net',0), a.get('treso_caisses_net1',0), indent=2)

    R("TOTAL III — TRÉSORERIE ACTIF",
      a.get('total_iii_brut',0), a.get('total_iii_net',0), a.get('total_iii_net1',0),
      is_total=True)

    R("TOTAL GÉNÉRAL ACTIF (I + II + III)",
      a.get('total_general_brut',0), a.get('total_general_net',0), a.get('total_general_net1',0),
      is_total=True)

    return row - 1


# ── Feuille 3 : Bilan Passif ───────────────────────────────────────────────────

def _write_passif(wb, passif: dict, exercice_fin: str) -> int:
    ws = wb.create_sheet("3 - Bilan Passif")
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, 'A', 52)
    _set_col_width(ws, 'B', 22)
    _set_col_width(ws, 'C', 22)

    p = passif
    row = 1

    _header_row(ws, row,
        ["BILAN PASSIF — MCN loi 9-88",
         f"Exercice {exercice_fin}", "Exercice Précédent"],
        height=22)
    row += 1

    def R(libelle, val_n, val_n1,
          is_total=False, is_section=False, is_subtotal=False, indent=0):
        nonlocal row
        _data_row(ws, row, libelle, val_n, val_n1,
                  is_total=is_total, is_section=is_section,
                  is_subtotal=is_subtotal, indent=indent)
        row += 1

    # ── FINANCEMENT PERMANENT ───────────────────────────────────────────────
    R("A — FINANCEMENT PERMANENT", 0, 0, is_section=True)

    R("Capitaux propres", p.get('cp_total_n',0), p.get('cp_total_n1',0), is_subtotal=True)
    R("  Capital social ou personnel", p.get('cp_capital_n',0), p.get('cp_capital_n1',0), indent=2)
    R("  Capital appelé", p.get('cp_capital_appele_n',0), p.get('cp_capital_appele_n1',0), indent=2)
    R("    dont versé", p.get('cp_capital_appele_n',0), p.get('cp_capital_appele_n1',0), indent=4)
    R("  Prime d'émission, de fusion, d'apport", 0, 0, indent=2)
    R("  Écarts de réévaluation", 0, 0, indent=2)
    R("  Réserves légales", p.get('cp_reserve_legale_n',0), p.get('cp_reserve_legale_n1',0), indent=2)
    R("  Autres réserves", p.get('cp_autres_reserves_n',0), p.get('cp_autres_reserves_n1',0), indent=2)
    R("  Reports à nouveau", p.get('cp_reports_nouveau_n',0), p.get('cp_reports_nouveau_n1',0), indent=2)
    R("  Résultats nets en instance d'affectation", 0, 0, indent=2)
    R("  Résultat net de l'exercice", p.get('cp_resultat_net_n',0), p.get('cp_resultat_net_n1',0), indent=2)
    R("  Total Capitaux Propres (A)", p.get('cp_total_n',0), p.get('cp_total_n1',0), is_subtotal=True)

    R("Capitaux propres assimilés (B)", p.get('cpa_total_n',0), p.get('cpa_total_n1',0), is_subtotal=True)
    R("  Subventions d'investissement", p.get('cpa_subventions_n',0), p.get('cpa_subventions_n1',0), indent=2)
    R("  Provisions réglementées", p.get('cpa_prov_reglementees_n',0), p.get('cpa_prov_reglementees_n1',0), indent=2)

    R("Dettes de financement (C)", p.get('dfi_total_n',0), p.get('dfi_total_n1',0), is_subtotal=True)
    R("  Emprunts obligataires", p.get('dfi_emprunts_oblig_n',0), p.get('dfi_emprunts_oblig_n1',0), indent=2)
    R("  Autres dettes de financement", p.get('dfi_autres_dettes_n',0), p.get('dfi_autres_dettes_n1',0), indent=2)

    R("Provisions durables pour risques et charges (D)",
      p.get('pdrc_total_n',0), p.get('pdrc_total_n1',0), is_subtotal=True)
    R("  Provisions pour risques", p.get('pdrc_risques_n',0), p.get('pdrc_risques_n1',0), indent=2)
    R("  Provisions pour charges", p.get('pdrc_charges_n',0), p.get('pdrc_charges_n1',0), indent=2)

    R("Écarts de conversion — Passif (E)", p.get('ecp_total_n',0), p.get('ecp_total_n1',0), is_subtotal=True)

    R("TOTAL I — FINANCEMENT PERMANENT", p.get('total_i_n',0), p.get('total_i_n1',0), is_total=True)

    # ── PASSIF CIRCULANT ────────────────────────────────────────────────────
    R("B — PASSIF CIRCULANT", 0, 0, is_section=True)

    R("Dettes du passif circulant (F)", p.get('dpc_total_n',0), p.get('dpc_total_n1',0), is_subtotal=True)
    R("  Fournisseurs et comptes rattachés", p.get('dpc_fournisseurs_n',0), p.get('dpc_fournisseurs_n1',0), indent=2)
    R("  Clients créditeurs, avances et acomptes", p.get('dpc_clients_credit_n',0), p.get('dpc_clients_credit_n1',0), indent=2)
    R("  Personnel", p.get('dpc_personnel_n',0), p.get('dpc_personnel_n1',0), indent=2)
    R("  Organismes sociaux", p.get('dpc_orga_sociaux_n',0), p.get('dpc_orga_sociaux_n1',0), indent=2)
    R("  État", p.get('dpc_etat_n',0), p.get('dpc_etat_n1',0), indent=2)
    R("  Comptes d'associés", p.get('dpc_comptes_assoc_n',0), p.get('dpc_comptes_assoc_n1',0), indent=2)
    R("  Autres créanciers", p.get('dpc_autres_n',0), p.get('dpc_autres_n1',0), indent=2)
    R("  Comptes de régularisation — Passif", p.get('dpc_regularisation_n',0), p.get('dpc_regularisation_n1',0), indent=2)

    R("Autres provisions pour risques et charges (G)",
      p.get('aprc_total_n',0), p.get('aprc_total_n1',0), is_subtotal=True)

    R("Écarts de conversion — Passif circulant (H)",
      p.get('ecpc_total_n',0), p.get('ecpc_total_n1',0), is_subtotal=True)

    R("TOTAL II — PASSIF CIRCULANT", p.get('total_ii_n',0), p.get('total_ii_n1',0), is_total=True)

    # ── TRÉSORERIE PASSIF ───────────────────────────────────────────────────
    R("C — TRÉSORERIE PASSIF", 0, 0, is_section=True)

    R("Crédit d'escompte", p.get('tp_credit_escompte_n',0), p.get('tp_credit_escompte_n1',0), indent=2)
    R("Crédit de trésorerie", p.get('tp_credit_tresorerie_n',0), p.get('tp_credit_tresorerie_n1',0), indent=2)
    R("Banques (soldes créditeurs)", p.get('tp_banques_n',0), p.get('tp_banques_n1',0), indent=2)

    R("TOTAL III — TRÉSORERIE PASSIF", p.get('total_iii_n',0), p.get('total_iii_n1',0), is_total=True)

    R("TOTAL GÉNÉRAL PASSIF (I + II + III)",
      p.get('total_general_n',0), p.get('total_general_n1',0), is_total=True)

    return row - 1


# ── Feuille 4 : CPC ────────────────────────────────────────────────────────────

def _write_cpc(wb, cpc: dict, exercice_fin: str) -> int:
    ws = wb.create_sheet("4 - CPC")
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, 'A', 52)
    _set_col_width(ws, 'B', 22)
    _set_col_width(ws, 'C', 22)

    c = cpc
    row = 1

    _header_row(ws, row,
        ["COMPTE DE PRODUITS ET CHARGES (Hors Taxes) — MCN loi 9-88",
         f"Exercice {exercice_fin}", "Exercice Précédent"],
        height=22)
    row += 1

    def R(libelle, val_n, val_n1,
          is_total=False, is_section=False, is_subtotal=False, indent=0):
        nonlocal row
        _data_row(ws, row, libelle, val_n, val_n1,
                  is_total=is_total, is_section=is_section,
                  is_subtotal=is_subtotal, indent=indent)
        row += 1

    # ── EXPLOITATION ─────────────────────────────────────────────────────────
    R("I — PRODUITS D'EXPLOITATION", 0, 0, is_section=True)
    R("  Ventes de marchandises (en l'état)",
      c.get('ventes_marchandises_n',0), c.get('ventes_marchandises_n1',0), indent=2)
    R("  Ventes de biens et services produits",
      c.get('ventes_biens_services_n',0), c.get('ventes_biens_services_n1',0), indent=2)
    R("  Chiffre d'affaires",
      c.get('chiffre_affaires_n',0), c.get('chiffre_affaires_n1',0), is_subtotal=True)
    R("  Variation de stocks de produits (+/-)",
      c.get('variation_stocks_produits_n',0), c.get('variation_stocks_produits_n1',0), indent=2)
    R("  Immobilisations produites par l'entreprise pour elle-même", 0, 0, indent=2)
    R("  Subventions d'exploitation", 0, 0, indent=2)
    R("  Autres produits d'exploitation",
      c.get('autres_produits_exploit_n',0), c.get('autres_produits_exploit_n1',0), indent=2)
    R("  Reprises d'exploitation ; transferts de charges",
      c.get('reprises_exploit_n',0), c.get('reprises_exploit_n1',0), indent=2)
    R("TOTAL I — Produits d'exploitation",
      c.get('total_produits_exploit_n',0), c.get('total_produits_exploit_n1',0), is_total=True)

    R("II — CHARGES D'EXPLOITATION", 0, 0, is_section=True)
    R("  Achats revendus de marchandises",
      c.get('achats_revendus_n',0), c.get('achats_revendus_n1',0), indent=2)
    R("  Achats consommés de matières et fournitures",
      c.get('achats_consommes_n',0), c.get('achats_consommes_n1',0), indent=2)
    R("  Autres charges externes",
      c.get('autres_charges_ext_n',0), c.get('autres_charges_ext_n1',0), indent=2)
    R("  Impôts et taxes",
      c.get('impots_taxes_n',0), c.get('impots_taxes_n1',0), indent=2)
    R("  Charges de personnel",
      c.get('charges_personnel_n',0), c.get('charges_personnel_n1',0), indent=2)
    R("  Autres charges d'exploitation",
      c.get('autres_charges_exploit_n',0), c.get('autres_charges_exploit_n1',0), indent=2)
    R("  Dotations d'exploitation",
      c.get('dot_exploitation_n',0), c.get('dot_exploitation_n1',0), indent=2)
    R("TOTAL II — Charges d'exploitation",
      c.get('total_charges_exploit_n',0), c.get('total_charges_exploit_n1',0), is_total=True)

    R("II — RÉSULTAT D'EXPLOITATION (I - II)",
      c.get('resultat_exploitation_n',0), c.get('resultat_exploitation_n1',0),
      is_total=True)

    # ── FINANCIER ────────────────────────────────────────────────────────────
    R("IV — PRODUITS FINANCIERS", 0, 0, is_section=True)
    R("  Produits des titres de participation et autres titres immobilisés",
      c.get('produits_titres_n',0), c.get('produits_titres_n1',0), indent=2)
    R("  Gains de change",
      c.get('gains_change_n',0), c.get('gains_change_n1',0), indent=2)
    R("  Intérêts et autres produits financiers",
      c.get('interets_produits_n',0), c.get('interets_produits_n1',0), indent=2)
    R("  Reprises financières ; transferts de charges",
      c.get('reprises_financieres_n',0), c.get('reprises_financieres_n1',0), indent=2)
    R("TOTAL IV — Produits financiers",
      c.get('total_produits_fin_n',0), c.get('total_produits_fin_n1',0), is_total=True)

    R("V — CHARGES FINANCIÈRES", 0, 0, is_section=True)
    R("  Charges d'intérêts",
      c.get('charges_interets_n',0), c.get('charges_interets_n1',0), indent=2)
    R("  Pertes de change",
      c.get('pertes_change_n',0), c.get('pertes_change_n1',0), indent=2)
    R("  Autres charges financières",
      c.get('autres_charges_fin_n',0), c.get('autres_charges_fin_n1',0), indent=2)
    R("  Dotations financières",
      c.get('dot_financieres_n',0), c.get('dot_financieres_n1',0), indent=2)
    R("TOTAL V — Charges financières",
      c.get('total_charges_fin_n',0), c.get('total_charges_fin_n1',0), is_total=True)

    R("VI — RÉSULTAT FINANCIER (IV - V)",
      c.get('resultat_financier_n',0), c.get('resultat_financier_n1',0), is_total=True)

    R("VII — RÉSULTAT COURANT (II + VI)",
      c.get('resultat_courant_n',0), c.get('resultat_courant_n1',0), is_total=True)

    # ── NON COURANT ──────────────────────────────────────────────────────────
    R("VIII — PRODUITS NON COURANTS", 0, 0, is_section=True)
    R("  Produits des cessions d'immobilisations",
      c.get('produits_cessions_n',0), c.get('produits_cessions_n1',0), indent=2)
    R("  Subventions d'équilibre", 0, 0, indent=2)
    R("  Reprises sur subventions d'investissement", 0, 0, indent=2)
    R("  Autres produits non courants",
      c.get('autres_produits_nc_n',0), c.get('autres_produits_nc_n1',0), indent=2)
    R("  Reprises non courantes ; transferts de charges",
      c.get('reprises_nc_n',0), c.get('reprises_nc_n1',0), indent=2)
    R("TOTAL VIII — Produits non courants",
      c.get('total_produits_nc_n',0), c.get('total_produits_nc_n1',0), is_total=True)

    R("IX — CHARGES NON COURANTES", 0, 0, is_section=True)
    R("  Valeurs nettes d'amortissement des immobilisations cédées",
      c.get('vna_immobilisations_n',0), c.get('vna_immobilisations_n1',0), indent=2)
    R("  Subventions accordées",
      c.get('subventions_accordees_n',0), c.get('subventions_accordees_n1',0), indent=2)
    R("  Autres charges non courantes",
      c.get('autres_charges_nc_n',0), c.get('autres_charges_nc_n1',0), indent=2)
    R("  Dotations non courantes aux amortissements et aux provisions",
      c.get('dot_nc_amort_prov_n',0), c.get('dot_nc_amort_prov_n1',0), indent=2)
    R("TOTAL IX — Charges non courantes",
      c.get('total_charges_nc_n',0), c.get('total_charges_nc_n1',0), is_total=True)

    R("X — RÉSULTAT NON COURANT (VIII - IX)",
      c.get('resultat_nc_n',0), c.get('resultat_nc_n1',0), is_total=True)

    R("XI — RÉSULTAT AVANT IMPÔTS (VII +/- X)",
      c.get('resultat_avant_impots_n',0), c.get('resultat_avant_impots_n1',0), is_total=True)

    R("XII — IMPÔTS SUR LES RÉSULTATS",
      c.get('impots_resultats_n',0), c.get('impots_resultats_n1',0), is_subtotal=True)

    R("XIII — RÉSULTAT NET (XI - XII)",
      c.get('resultat_net_n',0), c.get('resultat_net_n1',0), is_total=True)

    return row - 1


# ── Entrée publique ─────────────────────────────────────────────────────────────

def write(parsed: dict, output_path: str) -> dict:
    """
    Génère le fichier Excel structuré depuis les données parsées.
    Compatible avec l'interface attendue par app.py.
    """
    info    = parsed.get('info',   {})
    actif   = parsed.get('actif',  {})
    passif  = parsed.get('passif', {})
    cpc     = parsed.get('cpc',    {})
    exercice = info.get('exercice_fin', '')

    wb = openpyxl.Workbook()
    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    _write_identification(wb, info)
    n_actif  = _write_actif(wb, actif, exercice)
    n_passif = _write_passif(wb, passif, exercice)
    n_cpc    = _write_cpc(wb, cpc, exercice)

    # Freeze panes sur chaque feuille
    for ws in wb.worksheets:
        ws.freeze_panes = 'B2'

    wb.save(output_path)
    logger.info(f"SGTM Excel écrit : {output_path}")

    return {
        'format': 'SGTM',
        'rows':   n_actif + n_passif + n_cpc,
        'actif':  n_actif,
        'passif': n_passif,
        'cpc':    n_cpc,
    }
