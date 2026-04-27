"""
core/excel_writer.py
Génère l'Excel structuré à partir des données parsées (AMMC ou DGI).
Structure IDENTIQUE pour les deux formats :
  - 1 - Identification
  - 2 - Bilan Actif    : Label | Brut | Amort. | Net(N) | Net(N-1)
  - 3 - Bilan Passif   : Label | Exercice N | Exercice N-1
  - 4 - CPC            : N° | Label | Propres N | Exerc.Préc | Total N | Total N-1
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logger import get_logger

logger = get_logger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
C_DARK   = "1F4E79"
C_MED    = "2E75B6"
C_LIGHT  = "D6E4F0"
C_RESULT = "2E4057"
C_WHITE  = "FFFFFF"
C_GRAY   = "F5F7FA"
C_BORDER = "B8CCE4"
C_GOLD   = "FFF2CC"
NUM_FMT  = '#,##0.00;[Red]-#,##0.00;"-"'

# ── Styles ────────────────────────────────────────────────────────────────────

def _border():
    s = Side(style='thin', color=C_BORDER)
    return Border(top=s, bottom=s, left=s, right=s)

def _c(ws, r, c, v=None, bg=C_WHITE, fg="222222", bold=False,
       align="left", num_fmt=None, sz=9, wrap=True, indent=0):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", size=sz, bold=bold, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap, indent=indent)
    cell.border    = _border()
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def _row_colors(typ: str):
    if typ == 'total':   return C_DARK,   C_WHITE, True
    if typ == 'result':  return C_RESULT, C_WHITE, True
    if typ == 'section': return C_LIGHT,  C_DARK,  True
    return C_WHITE, "333333", False

# ── Bloc titre commun ─────────────────────────────────────────────────────────

def _title_block(ws, title: str, info: dict, n_cols: int) -> int:
    raison   = info.get('raison_sociale', '')
    if_num   = info.get('identifiant_fiscal', '')
    exercice = info.get('exercice_fin', '') or info.get('exercice', '')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    _c(ws, 1, 1, title, bg=C_DARK, fg=C_WHITE, bold=True, align='center', sz=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, n_cols-1))
    _c(ws, 2, 1, f"Raison sociale : {raison}", bg=C_LIGHT, fg=C_DARK, bold=True, sz=9, indent=1)
    _c(ws, 2, n_cols, f"IF : {if_num}", bg=C_LIGHT, fg=C_DARK, align='right', sz=9)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    _c(ws, 3, 1, f"Date de bilan : {exercice}", bg=C_GRAY, fg="555555", sz=9, indent=1)
    ws.row_dimensions[4].height = 4
    return 5

# ── Feuille Identification ────────────────────────────────────────────────────

def _write_ident(wb, info: dict):
    ws = wb.create_sheet("1 - Identification")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 54

    # Titre principal
    ws.merge_cells('A1:B1')
    _c(ws, 1, 1, "PIÈCES ANNEXES À LA DÉCLARATION FISCALE",
       bg=C_DARK, fg=C_WHITE, bold=True, align='center', sz=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:B2')
    _c(ws, 2, 1, "IMPÔTS SUR LES SOCIÉTÉS — Modèle Comptable Normal (loi 9-88)",
       bg=C_MED, fg=C_WHITE, align='center', sz=10)
    ws.row_dimensions[2].height = 18

    # Séparateur
    ws.merge_cells('A3:B3')
    _c(ws, 3, 1, "▌ Informations fiscales", bg=C_LIGHT, fg=C_DARK, bold=True, sz=9, indent=1)
    ws.row_dimensions[3].height = 16

    # Champs extraits du PDF
    fiscal_fields = [
        ("Raison sociale",       info.get('raison_sociale', '—')),
        ("Identifiant fiscal",   info.get('identifiant_fiscal', '—')),
        ("Taxe professionnelle", info.get('taxe_professionnelle', '—')),
        ("Adresse",              info.get('adresse', '—')),
        ("Date de bilan",        info.get('exercice_fin', '—') or info.get('exercice', '—')),
        ("Format PDF",           info.get('format', '—')),
    ]
    r = 4
    for lbl, val in fiscal_fields:
        ws.row_dimensions[r].height = 18
        _c(ws, r, 1, lbl, bg=C_LIGHT, fg=C_DARK, bold=True, sz=9, indent=1)
        _c(ws, r, 2, val, bg=C_WHITE,  fg="222222", sz=9, indent=1)
        r += 1

    # Séparateur section commerciale
    ws.merge_cells(f'A{r}:B{r}')
    _c(ws, r, 1, "▌ Informations commerciales", bg=C_LIGHT, fg=C_DARK, bold=True, sz=9, indent=1)
    ws.row_dimensions[r].height = 16
    r += 1

    # Champs commerciaux (centre d'affaires + macro-secteur)
    commercial_fields = [
        ("Centre d'affaires",    info.get('centre_affaires', '—')),
        ("Macro-secteur",        info.get('macro_secteur', '—')),
    ]
    for lbl, val in commercial_fields:
        ws.row_dimensions[r].height = 18
        _c(ws, r, 1, lbl, bg=C_GOLD, fg=C_DARK, bold=True, sz=9, indent=1)
        _c(ws, r, 2, val, bg=C_WHITE, fg="222222", sz=9, indent=1)
        r += 1

# ── Feuille Bilan Actif ───────────────────────────────────────────────────────

def _write_actif(wb, info: dict, template: list, value_map: dict):
    ws = wb.create_sheet("2 - Bilan Actif")
    ws.sheet_view.showGridLines = False
    n = 5
    r = _title_block(ws, "BILAN ACTIF", info, n)

    for ci, (h, w) in enumerate([
        ("DÉSIGNATION",         50),
        ("BRUT",                18),
        ("AMORT. & PROV.",      18),
        ("NET — EXERCICE N",    18),
        ("NET — EXERCICE N-1",  18),
    ], 1):
        _c(ws, r, ci, h, bg=C_MED, fg=C_WHITE, bold=True, align='center', sz=9, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 28
    ws.freeze_panes = f'A{r+1}'
    r += 1

    for ti, (key, label, typ) in enumerate(template):
        bg, fg, bold = _row_colors(typ)
        ws.row_dimensions[r].height = 15 if typ == 'normal' else 17
        indent = 1 if typ == 'normal' else 0
        _c(ws, r, 1, label, bg=bg, fg=fg, bold=bold, align='left', sz=9, indent=indent)
        vals = value_map.get(ti, [])
        for ci in range(4):
            v = vals[ci] if ci < len(vals) else None
            _c(ws, r, ci+2, 0 if v is None else v,
               bg=bg, fg=fg, bold=bold, align='right', sz=9, num_fmt=NUM_FMT)
        r += 1

    return r - 6

# ── Feuille Bilan Passif ──────────────────────────────────────────────────────

def _write_passif(wb, info: dict, template: list, value_map: dict):
    ws = wb.create_sheet("3 - Bilan Passif")
    ws.sheet_view.showGridLines = False
    n = 3
    r = _title_block(ws, "BILAN PASSIF", info, n)

    for ci, (h, w) in enumerate([
        ("DÉSIGNATION",  54),
        ("EXERCICE N",   20),
        ("EXERCICE N-1", 20),
    ], 1):
        _c(ws, r, ci, h, bg=C_MED, fg=C_WHITE, bold=True, align='center', sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 22
    ws.freeze_panes = f'A{r+1}'
    r += 1

    for ti, (key, label, typ) in enumerate(template):
        bg, fg, bold = _row_colors(typ)
        ws.row_dimensions[r].height = 15 if typ == 'normal' else 17
        indent = 1 if typ == 'normal' else 0
        _c(ws, r, 1, label, bg=bg, fg=fg, bold=bold, align='left', sz=9, indent=indent)
        vals = value_map.get(ti, [])
        for ci in range(2):
            v = vals[ci] if ci < len(vals) else None
            _c(ws, r, ci+2, 0 if v is None else v,
               bg=bg, fg=fg, bold=bold, align='right', sz=9, num_fmt=NUM_FMT)
        r += 1

    r2 = r + 1
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=3)
    c = ws.cell(r2, 1)
    c.value = "(1) Capital personnel débiteur.  (2) Bénéficiaire (+) / Déficitaire (−)."
    c.font  = Font(name="Arial", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return r - 6

# ── Feuille CPC ───────────────────────────────────────────────────────────────

_CPC_NUMS = [
    '',    '',    '',    '',    '',    '',    '',    '',    '',    'I',
    '',    '',    '',    '',    '',    '',    '',    '',    'II',  'III',
    '',    '',    '',    '',    '',    'IV',  '',    '',    '',    '',
    '',    'V',   'VI',  'VII', '',    '',    '',    '',    '',    '',
    'VIII','',    '',    '',    '',    '',    'IX',  'X',   'XI',  'XII',
    'XIII','XIV', 'XV',  'XVI',
]

def _write_cpc(wb, info: dict, template: list, value_map: dict):
    ws = wb.create_sheet("4 - CPC")
    ws.sheet_view.showGridLines = False
    n = 6
    r = _title_block(ws, "COMPTE DE PRODUITS ET CHARGES (Hors Taxes)", info, n)

    for ci, (h, w) in enumerate([
        ("N°",                     5),
        ("DÉSIGNATION",            42),
        ("PROPRES À\nL'EXERCICE",  18),
        ("EXERCICES\nPRÉCÉDENTS",  18),
        ("TOTAUX\nEXERCICE N",     18),
        ("TOTAUX\nEXERCICE N-1",   18),
    ], 1):
        _c(ws, r, ci, h, bg=C_MED, fg=C_WHITE, bold=True, align='center', sz=9, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 30
    ws.freeze_panes = f'A{r+1}'
    r += 1

    for ti, (key, label, typ) in enumerate(template):
        bg, fg, bold = _row_colors(typ)
        ws.row_dimensions[r].height = 15 if typ == 'normal' else 17
        indent = 1 if typ == 'normal' else 0
        num = _CPC_NUMS[ti] if ti < len(_CPC_NUMS) else ''
        _c(ws, r, 1, num or None, bg=bg,
           fg=fg if typ != 'normal' else C_MED,
           bold=True, align='center', sz=8)
        _c(ws, r, 2, label, bg=bg, fg=fg, bold=bold, align='left', sz=9, indent=indent)
        vals = value_map.get(ti, [])
        for ci in range(4):
            v = vals[ci] if ci < len(vals) else None
            _c(ws, r, ci+3, 0 if v is None else v,
               bg=bg, fg=fg, bold=bold, align='right', sz=9, num_fmt=NUM_FMT)
        r += 1

    r2 = r + 1
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=6)
    c = ws.cell(r2, 1)
    c.value = "(1) Stock final − Stock initial.  (2) Achats revendus = Achats − Variation de stock."
    c.font  = Font(name="Arial", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return r - 6

# ── Point d'entrée ────────────────────────────────────────────────────────────

def write(parsed: dict, output_path: str) -> dict:
    info      = parsed['info']
    templates = parsed['templates']
    fmt       = parsed.get('format', '?')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_ident(wb, info)
    n_a = _write_actif(wb,  info, templates['actif'],  parsed['actif'])
    n_p = _write_passif(wb, info, templates['passif'], parsed['passif'])
    n_c = _write_cpc(wb,    info, templates['cpc'],    parsed['cpc'])

    wb.save(output_path)
    total = n_a + n_p + n_c
    logger.info(f"Excel {fmt} : {total} lignes ({n_a}a/{n_p}p/{n_c}c) → {output_path}")

    return {'rows': total, 'actif': n_a, 'passif': n_p, 'cpc': n_c, 'format': fmt}
